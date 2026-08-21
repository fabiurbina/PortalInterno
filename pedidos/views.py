from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login,logout
from django.contrib.auth.decorators import login_required
from .omie_service import (listar_ops, consultar_produto,consultar_op,listar_locais_estoque,
consultar_pedido, extrair_numero_pedido, listar_lotes,listar_quarentena, listar_entradas_com_fornecedor, settings, buscar_cliente_cnpj,consultar_estrutura)
from django.contrib import messages
from django.utils import timezone
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse  
from django.core.paginator import Paginator
from django.core.mail import send_mail
import socket
from django.http import HttpResponse
import json
from .mysql_service import (
    datetime,
    inserir_inspecao,
    inserir_resultado_inspecao,
    salvar_conferencia_mysql,
    consultar_conferencias,
    salvar_controle_qualidade,
    salvar_observacao_op,
    salvar_apontamento,
    consultar_apontamentos,
    consultar_todos_pedidos,
    buscar_relatorio_mrp, 
    buscar_CRM, buscar_producao,
    buscar_posicao_estoque,
    salvar_setup,
    salvar_parada_sql,
    buscar_previsao_demanda,
    consulta_chao_fabrica,
    salvar_ordem_chao_fabrica
)
from django.core.cache import cache
from .status_service import interpretar_status
from collections import defaultdict
import pandas as pd
from io import BytesIO
from django.contrib.auth.models import User
import secrets
import string
from .email_service import enviar_email_boas_vindas
from .groq_service import gerar_relatorio_comercial, gerar_relatorio_producao
from .preparar_dados import preparar_dados_comercial, preparar_dados_producao
import markdown
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime



@login_required
def dashboard(request):

    return render(
        request,
        'dashboard.html'
    )
    
    
def home(request):

    ops = []

    ETAPAS = {
        "10": "Planejamento",
        "20": "Produzindo",
        "50": "Concluído",
        "60": "Armazenar/Expedir"
    }

    if request.method == 'POST':

        ultima_consulta = request.session.get('ultima_consulta')

        agora = timezone.now().timestamp()

        if ultima_consulta and (agora - ultima_consulta) < 3:

            messages.warning(
                request,
                'Aguarde 3 segundos antes de realizar uma nova consulta.'
            )

            return render(
                request,
                'home.html',
                {'ops': []}
            )

        request.session['ultima_consulta'] = agora

        etapa_filtro = request.POST.get('etapa')

        retorno = listar_ops()

        ops = retorno.get('cadastros', [])

        # adiciona o nome da etapa em todas as OPs
        produtos_cache = {}

        for op in ops:

            codigo = op['infAdicionais']['cEtapa']

            op['nome_etapa'] = ETAPAS.get(
                codigo,
                f"Etapa {codigo}"
            )

            cod_produto = op['identificacao']['nCodProduto']

            if cod_produto not in produtos_cache:

                produto = consultar_produto(cod_produto)

                produtos_cache[cod_produto] = produto.get(
                    'descricao',
                    str(cod_produto)
                )

            op['nome_produto'] = produtos_cache[cod_produto]

        # aplica o filtro se selecionado
        if etapa_filtro:

            ops = [
                op for op in ops
                if op['infAdicionais']['cEtapa'] == etapa_filtro
            ]

    return render(
        request,
        'home.html',
        {
            'ops': ops
        }
    )
    
    

def login_view(request):

    if request.method == 'POST':

        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(
            request,
            username=username,
            password=password
        )

        if user is not None:
            login(request, user)
            return redirect('/')

    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    return redirect('/login/')


def ficha_op(request, codigo_op):

    op = consultar_op(codigo_op)
    
    if 'identificacao' not in op:

        return render(
            request,
            'ficha_op.html',
            {
                'erro': op.get(
                    'faultstring',
                    'Erro ao consultar OP'
                )
            }
        )
        
    # Observação da OP
    observacao_op = op.get('observacoes', {}).get('cObs', '')

    # Busca descrição das matérias-primas
    produtos_cache = {}

    for item in op.get('itensDetalhes', []):

        codigo = item['nIdProdutoMalha']

        if codigo not in produtos_cache:

            produto_mp = consultar_produto(codigo)

            produtos_cache[codigo] = {
            'codigo': produto_mp.get('codigo', str(codigo)),
            'descricao': produto_mp.get('descricao', ''),
            'unidade': produto_mp.get('unidade', '')
}
        item['codigo_produto'] = produtos_cache[codigo]['codigo']
        item['descricao_produto'] = produtos_cache[codigo]['descricao']
        item['unidade_produto'] = produtos_cache[codigo]['unidade']

            
    locais = cache.get('locais_estoque')

    if not locais:

        locais = listar_locais_estoque()

        cache.set(
            'locais_estoque',
            locais,
            300
        )


    lotes = cache.get('lotes_omie')

    if not lotes:

        lotes = listar_lotes()

        cache.set(
            'lotes_omie',
            lotes,
            300
        )

    # Mapa de locais
    mapa_locais = {}

    for local in locais.get('locaisEncontrados', []):

        mapa_locais[
            local['codigo_local_estoque']
        ] = local['descricao']


    # Mapa de lotes
    mapa_lotes = {}
    
    

    for produto_lote in lotes.get('listaLotes', []):

        codigo_produto = produto_lote['ident']['nCodProd']

        if produto_lote.get('lotes'):

            lote = produto_lote['lotes'][0]

            mapa_lotes[codigo_produto] = {
                'lote': lote.get('cNumLote', ''),
                'validade': lote.get('dDataValidade', '')
            }

        else:

            mapa_lotes[codigo_produto] = {
                'lote': '',
                'validade': ''
            }
            
            
    #print(mapa_lotes)
    # Adiciona descrição do local + lote + validade
    for item in op.get('itensDetalhes', []):

        item['descricao_local'] = mapa_locais.get(
            item['codigo_local_estoque'],
            str(item['codigo_local_estoque'])
        )

        codigo = item['nIdProdutoMalha']

        item['lote'] = mapa_lotes.get(
            codigo,
            {}
        ).get(
            'lote',
            ''
        )

        item['validade'] = mapa_lotes.get(
            codigo,
            {}
        ).get(
            'validade',
            ''
        )
        

    mapa_locais = {}

    for local in locais.get('locaisEncontrados', []):

        mapa_locais[
            local['codigo_local_estoque']
        ] = local['descricao']


    for item in op.get('itensDetalhes', []):

        item['descricao_local'] = mapa_locais.get(
            item['codigo_local_estoque'],
            str(item['codigo_local_estoque'])
        )
        
        
    materias_primas = []
    embalagens = []

    for item in op.get('itensDetalhes', []):

        if item['descricao_local'] == 'Estoque Matéria Prima':
            materias_primas.append(item)

        elif item['descricao_local'] == 'Estoque de Embalagens':
            embalagens.append(item)


    # Calcula o peso total da batida
    peso_batida_kg = 0
    quantidade_capsulas = 0

    for item in materias_primas:

        unidade = item.get('unidade_produto', '').upper()
        quantidade = float(item.get('nQtde', 0) or 0)
        descricao = item.get('descricao_produto', '').upper()

        if unidade == 'KG':
            peso_batida_kg += quantidade

        elif unidade == 'UN' and 'CAPSULA' in descricao:
            quantidade_capsulas += quantidade


    # Calcula o peso teórico do conteúdo por cápsula em mg
    peso_por_capsula_mg = 0

    if quantidade_capsulas > 0:
        peso_por_capsula_mg = (
            peso_batida_kg * 1_000_000
        ) / quantidade_capsulas


    print("⚖️ PESO DA BATIDA:", peso_batida_kg, "KG")
    print("💊 QUANTIDADE DE CÁPSULAS:", quantidade_capsulas, "UN")
    print("📏 PESO TEÓRICO POR CÁPSULA:", peso_por_capsula_mg, "MG")
            
            

    produto = consultar_produto(
    op['identificacao']['nCodProduto']
    )

    codigo_pa = op['identificacao']['nCodProduto']
    print("Código PA:", codigo_pa)
    print("Código PA:", codigo_pa)
    estrutura = consultar_estrutura(codigo_pa)

    observacao_estrutura = (
        estrutura.get("observacoes", {})
                .get("obsRelevantes", "")
    )
    
    codigo_op = op['identificacao']['cNumOP']

    conferencias = consultar_conferencias(
        codigo_op
    )
    
    apontamentos = consultar_apontamentos(
    codigo_op)
    
    #print(apontamentos)
    
    mapa_apontamentos = {}

    for apontamento in apontamentos:

        mapa_apontamentos[
            apontamento['etapa']
        ] = apontamento

    mapa_conferencias = {}

    for conferencia in conferencias:

        mapa_conferencias[
            str(conferencia['codigo_produto'])
        ] = conferencia

    for item in op.get('itensDetalhes', []):
        

        item['conferido'] = (
            str(int(item['codigo_produto']))
            in mapa_conferencias
        )

    # Lote do Produto Acabado
    lote_pa = mapa_lotes.get(
        codigo_pa,
        {}
    ).get(
        'lote',
        ''
    )

    validade_pa = mapa_lotes.get(
        codigo_pa,
        {}
    ).get(
        'validade',
        ''
    )

    obs = op['observacoes']['cObs']

    #numero_pedido = extrair_numero_pedido(obs)

    #pedido = None

    #if numero_pedido:
        #pedido = consultar_pedido(numero_pedido)

    nome_produto = produto.get(
        'descricao',
        op['identificacao']['nCodProduto']
    )

    ETAPAS = {
            "10": "Planejamento",
            "20": "Produzindo",
            "50": "Concluído",
            "60": "Armazenar/Expedir"
        }

    nome_etapa = ETAPAS.get(
        op['infAdicionais']['cEtapa'],
        op['infAdicionais']['cEtapa']
    )
    
    ETAPAS_PRODUCAO = [
        "FRACIONAMENTO",
        "MISTURA",
        "ENVASE",
        "EXPEDICAO"
    ]
    
    MAQUINAS = [
        "Linha de Pós 01",
        "Linha de Cápsulas 01",]

    return render(
        request,
        'ficha_op.html',
        {
            'op': op,
            'observacao_op': observacao_op,
            'nome_produto': nome_produto,
            'nome_etapa': nome_etapa,
            'materias_primas': materias_primas,
            'embalagens': embalagens,
            'pedido': None,
            'numero_pedido': None,
            'lote_pa': lote_pa,
            'validade_pa': validade_pa,
            'etapas_producao': ETAPAS_PRODUCAO,
            'apontamentos': mapa_apontamentos,
            'maquinas': MAQUINAS,
             # Observações da estrutura
            'observacao_estrutura': observacao_estrutura,
        
            # Dados calculados da produção
            'peso_batida_kg': peso_batida_kg,
            'quantidade_capsulas': quantidade_capsulas,
            'peso_por_capsula_mg': peso_por_capsula_mg,
        }
    )
        
@csrf_exempt
def salvar_conferencia(request):

    if request.method == "POST":

        dados = json.loads(request.body)
        
        from datetime import datetime

        validade = None

        if dados['validade']:
            validade = datetime.strptime(
                dados['validade'],
                '%d/%m/%Y'
            ).date()
            
        validade_pa = None

        if dados['validade_pa']:
            validade_pa = datetime.strptime(
                dados['validade_pa'],
                '%d/%m/%Y'
            ).date()

        salvar_conferencia_mysql(
            codigo_op=dados['op'],
            codigo_produto=dados['produto'],
            descricao_produto=dados['descricao'],
            tipo=dados['tipo'],
            lote=dados['lote'],
            validade=validade,
            conferido=dados['conferido'],
            usuario=request.user.username,
            lote_pa = dados['lote_pa'],
            validade_pa = validade_pa,
            observacao = dados.get('observacao',''),
            limpeza_sala=dados.get('limpeza_sala'),
            limpeza_equipamento=dados.get('limpeza_equipamento')
            
        )

        print("Conferência salva!")

        return JsonResponse({
            "sucesso": True
        })

    return JsonResponse({
        "sucesso": False
    })
    
@csrf_exempt
def salvar_observacao(request):

    if request.method == "POST":

        dados = json.loads(request.body)

        salvar_observacao_op(
            codigo_op=dados['op'],
            observacao=dados['observacao'],
            usuario=request.user.username
        )

        return JsonResponse({
            "sucesso": True
        })

    return JsonResponse({
        "sucesso": False
    })
    

@csrf_exempt
def salvar_apontamento_view(request):

    try:

        if request.method == "POST":

            print("=" * 50)

            for campo, valor in request.POST.items():
                print(f"{campo} = {valor}")

            print("=" * 50)
            
            # ==========================================================
            # SETUP
            # ==========================================================

            hora_inicio_setup = request.POST.get("hora_inicio_setup")

            if hora_inicio_setup:

                salvar_setup(

                    codigo_op=request.POST.get("codigo_op"),

                    idMaquina=int(request.POST.get("id_maquina")),

                    hora_inicio=hora_inicio_setup,

                    hora_fim=request.POST.get("hora_fim_setup"),

                    usuario=request.user.username,

                    observacao=request.POST.get("observacao_setup")

                )
                
            # ==============================  
            # PARADAS DA OP
            # ==============================

            tipos_parada = request.POST.getlist("tipo_parada[]")
            motivos_parada = request.POST.getlist("motivo_parada[]")
            inicios_parada = request.POST.getlist("inicio_parada[]")
            fins_parada = request.POST.getlist("fim_parada[]")
            observacoes_parada = request.POST.getlist("observacao_parada[]")


            for tipo, motivo, inicio, fim, observacao in zip(
                tipos_parada,
                motivos_parada,
                inicios_parada,
                fins_parada,
                observacoes_parada
            ):

                if inicio:

                    salvar_parada_sql(
                        codigo_op=request.POST.get("codigo_op"),
                        tipo_parada=tipo,
                        motivo=motivo,
                        data_inicio=inicio,
                        data_fim=fim,
                        usuario=request.user.username,
                        observacao=observacao
                    )
            
            #Fracionamento    
            hora_inicio = request.POST.get("hora_inicio_fracionamento")
            if hora_inicio:
                
                salvar_apontamento(
                    
                    #idMaquina = 1 if request.POST.get("id_maquina") == "1" else 2,

                    codigo_op=request.POST.get("codigo_op"),

                    etapa="Fracionamento",

                    equipamento_limpo=True,

                    area_limpa=True,

                    mp_conferidas=True,

                    data_inicio=request.POST.get("hora_inicio_fracionamento"),

                    data_fim=request.POST.get("hora_fim_fracionamento"),

                    medida_prevista=request.POST.get("peso_previsto_fracionamento"),

                    medida_real=request.POST.get("peso_fracionado"),
                    
                    medida_perdas=0,

                    unidade="KG",

                    usuario=request.user.username,

                    observacao=request.POST.get("observacao_fracionamento")

                )
                
            #mistura
            hora_inicio = request.POST.get("hora_inicio_mistura")     
            if hora_inicio:

                salvar_apontamento(
                    
                   #idMaquina = 1 if request.POST.get("id_maquina") == "1" else 2,
                    
                    codigo_op=request.POST.get("codigo_op"),

                    etapa="Mistura",

                    equipamento_limpo=True,

                    area_limpa=True,

                    mp_conferidas=True,

                    data_inicio=request.POST.get("hora_inicio_mistura"),

                    data_fim=request.POST.get("hora_fim_mistura"),

                    medida_prevista=request.POST.get("peso_previsto_mistura"),

                    medida_real=request.POST.get("peso_obtido_mistura"),
                    
                    medida_perdas=0,

                    unidade="KG",

                    usuario=request.user.username,

                    observacao=request.POST.get("observacao_Mistura")

                )
                
            #Qualidade
            hora_inicio = request.POST.get("hora_inicio_qualidade")
            if hora_inicio:
                
            
                salvar_apontamento(
                    
                   #idMaquina = 1 if request.POST.get("id_maquina") == "1" else 2,

                    codigo_op=request.POST.get("codigo_op"),

                    etapa="Qualidade",

                    equipamento_limpo=True,

                    area_limpa=True,

                    mp_conferidas=True,
                    
                    medida_perdas=0,

                    data_inicio=request.POST.get("hora_inicio_qualidade"),

                    data_fim=request.POST.get("hora_fim_qualidade"),

                    usuario=request.user.username,

                    observacao=request.POST.get("observacao_qualidade")

                )
                
                
            #encapsulamento
            hora_inicio = request.POST.get("hora_inicio_encapsulamento")
            if hora_inicio:
                salvar_apontamento(
                    
                   #idMaquina = 1 if request.POST.get("id_maquina") == "1" else 2,

                    codigo_op=request.POST.get("codigo_op"),

                    etapa="encapsulamento",

                    equipamento_limpo=True,

                    area_limpa=True,

                    mp_conferidas=True,

                    data_inicio=request.POST.get("hora_inicio_encapsulamento"),

                    data_fim=request.POST.get("hora_fim_encapsulamento"),

                    medida_prevista=request.POST.get("quantidade_produzida_encapsulamento"),
                    
                    medida_perdas=0,

                    unidade="UN",

                    usuario=request.user.username,

                    observacao=request.POST.get("observacao_Encapsulamento")

                )
            
            #Envase           
            hora_inicio = request.POST.get("hora_inicio_envase")
            if hora_inicio:
                
                salvar_apontamento(
                    
                   #idMaquina = 1 if request.POST.get("id_maquina") == "1" else 2,

                    codigo_op=request.POST.get("codigo_op"),

                    etapa="Envase",

                    equipamento_limpo=True,

                    area_limpa=True,

                    mp_conferidas=True,

                    data_inicio=request.POST.get("hora_inicio_envase"),

                    data_fim=request.POST.get("hora_fim_envase"),

                    medida_prevista=request.POST.get("quantidade_produzida_envase"),

                    medida_real=request.POST.get("quantidade_expedida_envase"),

                    medida_perdas=0,

                    unidade="UN",

                    usuario=request.user.username,

                    observacao=request.POST.get("observacao_Envase")

                )
            

            return JsonResponse({
                "sucesso": True
            })

    except Exception as e:

        print(e)  # ou logger.exception(e)

        return JsonResponse({
            "sucesso": False,
            "erro": "Ocorreu um erro ao registrar o apontamento."
        })
        
        
        

    
    
def qualidade_home(request):
    

    return render(
        request,
        'qualidade_home.html'
    )
    


@login_required
def qualidade_home(request):

    data_inicial = request.GET.get(
        'data_inicial'
    )

    data_final = request.GET.get(
        'data_final'
    )

    entradas = []
    
    pendentes = 0
    aprovados = 0
    reprovados = 0

    if data_inicial and data_final:

        data_inicial = (
            data_inicial
            .split('-')
        )

        data_final = (
            data_final
            .split('-')
        )

        data_inicial = (
            f"{data_inicial[2]}/"
            f"{data_inicial[1]}/"
            f"{data_inicial[0]}"
        )

        data_final = (
            f"{data_final[2]}/"
            f"{data_final[1]}/"
            f"{data_final[0]}"
        )

        entradas = listar_entradas_com_fornecedor(
            data_inicial,
            data_final
        )
        
        pendentes = sum(
            1 for e in entradas
            if e["status"] == "Aguardando"
        )

        aprovados = sum(
            1 for e in entradas
            if e["status"] == "Aprovado"
        )

        reprovados = sum(
            1 for e in entradas
            if e["status"] == "Reprovado"
        )

    return render(
        request,
        'qualidade_home.html',
        {
            'entradas': entradas,
            'pendentes': pendentes,
            'aprovados': aprovados,
            'reprovados': reprovados,
        }
    )
    
     
@login_required
def estoque_home(request):

    ops = []

    ETAPAS = {
            "10": "Planejamento",
            "20": "Produzindo",
            "50": "Concluído",
            "60": "Armazenar/Expedir"
        }

    if request.method == 'POST':

        etapa_filtro = request.POST.get('etapa')

        retorno = listar_ops()

        ops = retorno.get('cadastros', [])

        produtos_cache = {}

        for op in ops:

            codigo = op['infAdicionais']['cEtapa']

            op['nome_etapa'] = ETAPAS.get(
                codigo,
                f"Etapa {codigo}"
            )

            cod_produto = op['identificacao']['nCodProduto']

            if cod_produto not in produtos_cache:

                produto = consultar_produto(cod_produto)

                produtos_cache[cod_produto] = produto.get(
                    'descricao',
                    str(cod_produto)
                )
                

            op['nome_produto'] = produtos_cache[cod_produto]
            

        if etapa_filtro:

            ops = [
                op for op in ops
                if op['infAdicionais']['cEtapa'] == etapa_filtro
            ]

    return render(
        request,
        'estoque_home.html',
        {
            'ops': ops
        }
    )
    
    
    
 
@login_required
def ficha_logistica(request, codigo_op):

    op = consultar_op(codigo_op)

    if 'identificacao' not in op:

        return render(
            request,
            'ficha_logistica.html',
            {
                'erro': 'Erro ao consultar OP'
            }
        )

    # Busca descrição dos produtos
    produtos_cache = {}

    for item in op.get('itensDetalhes', []):

        codigo = item['nIdProdutoMalha']

        if codigo not in produtos_cache:

            produto_mp = consultar_produto(codigo)

            produtos_cache[codigo] = {
                'codigo': produto_mp.get('codigo', str(codigo)),
                'descricao': produto_mp.get('descricao', ''),
                'familia': produto_mp.get('descricao_familia', '')
            }

        item['codigo_produto'] = produtos_cache[codigo]['codigo']
        item['descricao_produto'] = produtos_cache[codigo]['descricao']
        item['familia_produto'] = produtos_cache[codigo]['familia']
    print(item['descricao_produto'])
    print(item['familia_produto'])
    print('----------------')

    # Locais de estoque
    locais = listar_locais_estoque()

    mapa_locais = {}

    for local in locais.get('locaisEncontrados', []):

        mapa_locais[
            local['codigo_local_estoque']
        ] = local['descricao']

    # LOTES
    lotes = listar_lotes()

    mapa_lotes = {}

    for produto_lote in lotes.get('listaLotes', []):

        codigo_produto = produto_lote['ident']['nCodProd']

        if produto_lote.get('lotes'):

            lote = produto_lote['lotes'][0]

            mapa_lotes[codigo_produto] = {
                'lote': lote.get('cNumLote', ''),
                'validade': lote.get('dDataValidade', '')
            }

        else:

            mapa_lotes[codigo_produto] = {
                'lote': '',
                'validade': ''
            }
            

    materias_primas = []
    embalagens = []

    for item in op.get('itensDetalhes', []):

        item['descricao_local'] = mapa_locais.get(
            item['codigo_local_estoque'],
            str(item['codigo_local_estoque'])
        )

        codigo = item['nIdProdutoMalha']

        item['lote'] = mapa_lotes.get(
            codigo,
            {}
        ).get(
            'lote',
            ''
        )

        item['validade'] = mapa_lotes.get(
            codigo,
            {}
        ).get(
            'validade',
            ''
        )

        if item['descricao_local'] == 'Estoque Matéria Prima':

            materias_primas.append(item)

        elif item['descricao_local'] == 'Estoque de Embalagens':

            embalagens.append(item)

    return render(
        request,
        'ficha_logistica.html',
        {
            'op': op,
            'materias_primas': materias_primas,
            'embalagens': embalagens,
            'lote_pa': '',
            'validade_pa': '',
            'etapas_producao': []
        }
    )
    
@login_required
def ficha_qualidade(request, codigo_produto,codigo_op):
    
    
    op = consultar_op(codigo_op)
    

@login_required
def qualidade_inspecao(request, cod_prod):

    data_inicial = request.GET.get("data_inicial")
    data_final = request.GET.get("data_final")
    lote = request.GET.get("lote")

    # Garante que as datas existam
    if not data_inicial or not data_final:
        return redirect("qualidade_home")

    # Converte YYYY-MM-DD para DD/MM/YYYY
    partes = data_inicial.split("-")
    data_inicial_formatada = (
        f"{partes[2]}/"
        f"{partes[1]}/"
        f"{partes[0]}"
    )

    partes = data_final.split("-")
    data_final_formatada = (
        f"{partes[2]}/"
        f"{partes[1]}/"
        f"{partes[0]}"
    )

    entradas = listar_entradas_com_fornecedor(
        data_inicial_formatada,
        data_final_formatada
    )

    item = None

    # Produto + LOTE selecionado
    for entrada in entradas:
        if entrada["cod_prod"] == cod_prod:
            if entrada.get("lote") == lote:
                item = entrada
                break

    if item is None:
        return render(
            request,
            "qualidade_ficha.html",
            {
                "item": {
                    "cod_prod": cod_prod,
                    "descricao": "",
                    "fornecedor": "",
                    "codigo_fornecedor": "",
                    "quantidade": "",
                    "data": "",
                    "lote": lote or "",
                    "fabricacao": "",
                    "validade": ""
                },
                "erro": "Não foi possível localizar o lote. Refaça a pesquisa."
            }
        )

    return render(
        request,
        "qualidade_ficha.html",
        {
            "item": item,
            "data_inicial": data_inicial,
            "data_final": data_final
        }
    )
    
    
@login_required
def salvar_inspecao(request):

    if request.method != "POST":

        return redirect("qualidade_home")

    codigo_produto = request.POST.get("codigo_produto")
    codigo_fornecedor = request.POST.get("codigo_fornecedor")
    qtdd = request.POST.get("qtdd")
    data_recebimento = request.POST.get("data_recebimento")
    if data_recebimento:
        data_recebimento = datetime.strptime(
            data_recebimento, "%d/%m/%Y"
    ).date()
    lote = request.POST.get("lote")
    data_fabricacao = request.POST.get("data_fabricacao")
    if data_fabricacao:
        data_fabricacao = datetime.strptime(
            data_fabricacao, "%d/%m/%Y"
        ).date()
    data_validade = request.POST.get("data_validade")
    if data_validade:
        data_validade = datetime.strptime(
            data_validade, "%d/%m/%Y"
        ).date()
    data = request.POST.get("data")
    hora_inicio = request.POST.get("hora_inicio")
    hora_atual = datetime.now().strftime("%H:%M")
    if hora_inicio > hora_atual:
        messages.error(request, "A hora de início não pode ser maior que a hora atual.")
        return redirect(
            f"/qualidade/inspecao/{codigo_produto}/"
            f"?data_inicial={request.POST.get('data_inicial')}"
            f"&data_final={request.POST.get('data_final')}"
        )
    hora_fim = request.POST.get("hora_fim")
    if hora_fim < hora_inicio:
        messages.error(
            request,
            "A hora de término não pode ser menor que a hora de início.")
        return redirect(
        f"/qualidade/inspecao/{codigo_produto}/"
        f"?data_inicial={request.POST.get('data_inicial')}"
        f"&data_final={request.POST.get('data_final')}"
    )
    parecer = request.POST.get("parecer")
    motivo = request.POST.get("motivo")
    observacoes = request.POST.get("observacoes")
    responsavel = request.POST.get("responsavel")

    id_inspecao = inserir_inspecao(

        codigo_for="FOR-020-000",

        revisao="00",

        data_vigente="2026-05-27",

        codigo_produto=codigo_produto,

        codigo_fornecedor=codigo_fornecedor,

        qtdd=qtdd,

        data_recebimento=data_recebimento,

        status="Concluído",

        lote=lote,

        data_fabricacao=data_fabricacao,

        data_validade=data_validade,

        data=data,

        hora_inicio=hora_inicio,

        hora_fim=hora_fim,

        parecer=parecer,

        motivo=motivo,

        observacoes=observacoes,

        responsavel=responsavel

    )
    
    parametros = [

        (1, "", request.POST.get("embalagem"), None),

        (2, "", request.POST.get("umidade_visual"), None),

        (3, "", request.POST.get("avarias"), None),

        (4, "", request.POST.get("identificacao"), None),

        (5, "", request.POST.get("lacre"), None),

        (6,
        request.POST.get("aspecto"),
        request.POST.get("aspecto_conf"),
        None),

        (7,
        request.POST.get("cor"),
        request.POST.get("cor_conf"),
        None),

        (8,
        request.POST.get("odor"),
        request.POST.get("odor_conf"),
        None),

        (9,
        request.POST.get("materiais"),
        request.POST.get("materiais_conf"),
        None),

        (10,
        request.POST.get("densidade"),
        request.POST.get("densidade_conf"),
        None),

        (11,
        request.POST.get("umidade"),
        request.POST.get("umidade_conf"),
        None),

        (12,
        request.POST.get("solubilidade"),
        request.POST.get("solubilidade_conf"),
        None),

        (13,
        request.POST.get("ph"),
        request.POST.get("ph_conf"),
        None),

    ]
    
    for parametro in parametros:

        inserir_resultado_inspecao(

            id_inspecao=id_inspecao,

            id_parametro=parametro[0],

            resultado=parametro[1],

            conforme=parametro[2],

            observacao=parametro[3]

        )

    print(f"Inspeção de qualidade salva com ID: {id_inspecao}")

    return redirect(
    f"/qualidade/inspecao/{codigo_produto}/"
    f"?data_inicial={request.POST.get('data_inicial')}"
    f"&data_final={request.POST.get('data_final')}"
)
    
    
@login_required
def pedidos(request):

    numero_pedido = request.GET.get("pedido")
    
    status_filtro = request.GET.get("status")

    lista_pedidos = consultar_todos_pedidos()

    pedidos_processados = []

    for pedido in lista_pedidos:

        status = interpretar_status(
            pedido["descricao_etapa"],
            pedido["status_producao"],
        )

        pedido["titulo"] = status["titulo"]
        pedido["descricao"] = status["descricao"]
        pedido["cor"] = status["cor"]

        pedidos_processados.append(pedido)

    # FILTRO PELO NÚMERO DO PEDIDO
    if numero_pedido:

        pedidos_processados = [
            pedido
            for pedido in pedidos_processados
            if str(pedido["numero_pedido"]) == numero_pedido
            
        ]
        
     # FILTRO POR STATUS

    if status_filtro:

        pedidos_processados = [
            pedido
            for pedido in pedidos_processados
            if pedido["titulo"] == status_filtro
        ]

    # PAGINAÇÃO
    paginator = Paginator(
        pedidos_processados,
        20
    )

    pagina = request.GET.get("page")

    pedidos_processados = paginator.get_page(
        pagina
    )

    return render(
        request,
        "pedidos.html",
        {
            "pedidos": pedidos_processados,
            "status_selecionado": status_filtro,
            "pedido_pesquisado": numero_pedido,
        }
    )
    
    
def relatorio_mrp_view(request):

    dados_mrp = buscar_relatorio_mrp()

    pedidos_agrupados = defaultdict(list)

    for registro in dados_mrp:
        pedidos_agrupados[registro["numero_pedido"]].append(registro)

    pedidos = []

    for numero_pedido, registros_pedido in pedidos_agrupados.items():

        produtos_agrupados = defaultdict(list)

        for registro in registros_pedido:

            chave_produto = (
                registro.get("codigo"),
                registro.get("descricao")
            )

            produtos_agrupados[chave_produto].append(registro)

        produtos = []

        for (codigo_produto, descricao_produto), itens_produto in produtos_agrupados.items():

            produtos.append({

                "codigo": codigo_produto,

                "descricao": descricao_produto,

                "quantidade": itens_produto[0]["quantidade"],

                "unidade": itens_produto[0].get("unidade"),

                "numero_op": ", ".join(sorted(set(
                    str(item["numero_op"])
                    for item in itens_produto
                    if item.get("numero_op")
                ))),

                "quantidade_componentes": len(itens_produto),

                "itens": itens_produto

            })

        pedidos.append({

            "razao_social": registros_pedido[0].get("razao_social"),

            "numero_pedido": numero_pedido,

            "data_previsao": registros_pedido[0].get("data_previsao"),

            "quantidade_ops": len(set(
                item.get("numero_op")
                for item in registros_pedido
                if item.get("numero_op")
            )),

            "quantidade_componentes": len(registros_pedido),

            "produtos": produtos

        })

    context = {
        "pedidos": pedidos
    }

    return render(
        request,
        "relatorio_mrp.html",
        context
    )
    
def exportar_mrp_excel(request):

    dados_mrp = buscar_relatorio_mrp()

    df = pd.DataFrame(dados_mrp)

    # Converte colunas numéricas para número
    colunas_numericas = [
        "quantidade",
        "necessidade_componente",
        "estoque_atual",
        "saldo_provisionado"
    ]

    for coluna in colunas_numericas:
        if coluna in df.columns:
            df[coluna] = pd.to_numeric(
                df[coluna],
                errors="coerce"
            )

    # Ordenação
    df = df.sort_values(
        by=[
            "data_previsao",
            "numero_pedido",
            "codigo"
        ],
        ascending=[True, True, True]
    )

    # Ordem das colunas
    df = df[
        [
            "codigo_produto_malha",
            "descricao_produto_malha",
            "data_previsao",
            "numero_pedido",
            "razao_social",
            "codigo",
            "descricao",
            "quantidade",
            "unidade",
            "numero_op",
            "necessidade_componente",
            "estoque_atual",
            "saldo_provisionado"
        ]
    ]

    # Renomeia cabeçalhos
    df.rename(columns={
        "codigo_produto_malha": "Código Componente",
        "descricao_produto_malha": "Componente",
        "data_previsao": "Data Prevista",
        "numero_pedido": "Pedido",
        "razao_social": "Cliente",
        "codigo": "Código Produto",
        "descricao": "Produto",
        "quantidade": "Quantidade",
        "unidade": "Unidade",
        "numero_op": "OP",
        "necessidade_componente": "Necessidade",
        "estoque_atual": "Estoque Atual",
        "saldo_provisionado": "Saldo Projetado"
    }, inplace=True)

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        df.to_excel(
            writer,
            index=False,
            sheet_name="MRP"
        )

        worksheet = writer.sheets["MRP"]

        # Formatação numérica
        colunas_decimais = [
            "Quantidade",
            "Necessidade",
            "Estoque Atual",
            "Saldo Projetado"
        ]

        for coluna in worksheet.iter_cols():

            nome_coluna = coluna[0].value

            if nome_coluna in colunas_decimais:

                for cell in coluna[1:]:

                    if isinstance(cell.value, (int, float)):

                        # Até 2 casas decimais
                        # sem zeros desnecessários
                        cell.number_format = "#,##0.##"

        # Ajusta automaticamente a largura das colunas
        for coluna in worksheet.columns:

            maior = 0

            for cell in coluna:

                if cell.value is not None:
                    maior = max(
                        maior,
                        len(str(cell.value))
                    )

            letra = coluna[0].column_letter

            worksheet.column_dimensions[
                letra
            ].width = maior + 3

        # Congela a primeira linha
        worksheet.freeze_panes = "A2"

    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="relatorio_mrp.xlsx"'
    )

    return response


from django.http import HttpResponse

def teste_socket(request):
    try:
        with socket.create_connection(("smtp.gmail.com", 465), timeout=10):
            return HttpResponse("Gmail OK")
    except Exception as e:
        return HttpResponse(f"{type(e).__name__}: {e}")


def password_reset_view(request):

    if request.method == "POST":
        return redirect("password_reset_done")

    return render(request, "password_reset.html")

import re   

def limpar_documento(documento):
    if not documento:
        return ""
    return re.sub(r"\D", "", documento)


def gerar_senha(tamanho=10):

    caracteres = (
        string.ascii_letters +
        string.digits +
        "!@#$%&*"
    )

    return "".join(
        secrets.choice(caracteres)
        for _ in range(tamanho)
    )


@login_required
def criar_acesso_cliente(request):

    cliente = None

    if request.method == "POST":

        if "buscar" in request.POST:

            cnpj = limpar_documento(request.POST.get("cnpj"))

            cliente = buscar_cliente_cnpj(cnpj)

        elif "criar" in request.POST:

            cnpj = limpar_documento(request.POST.get("cnpj"))
            email = request.POST.get("email")

            cliente = buscar_cliente_cnpj(cnpj)

            if User.objects.filter(username=cnpj).exists():

                messages.error(
                    request,
                    "Este cliente já possui acesso."
                )

            else:

                senha = gerar_senha()

                User.objects.create_user(
                    username=cnpj,
                    email=email,
                    password=senha
                )
                
                

                cliente["cnpj_cpf"] = cnpj

                try:
                    enviar_email_boas_vindas(
                        cliente,
                        email,
                        senha
                    )

                    messages.success(
                        request,
                        "Usuário criado com sucesso! As credenciais foram enviadas por e-mail."
                    )

                except Exception as e:

                    messages.warning(
                        request,
                        f"Usuário criado, porém ocorreu um erro ao enviar o e-mail: {e}"
                    )

    return render(
        request,
        "criar_acesso_cliente.html",
        {
            "cliente": cliente
        }
    )
    
    
def indicadores_view(request):
    return render(request, "indicadores.html")
    
    
def analise_comercial(request):

    registros = buscar_CRM()

    dados = preparar_dados_comercial(registros)

    analise = gerar_relatorio_comercial(dados)

    html = markdown.markdown(
        analise,
        extensions=["tables", "fenced_code"]
    )

    return HttpResponse(html)


def analise_producao(request):

    registros = buscar_producao()

    dados = preparar_dados_producao(registros)

    analise = gerar_relatorio_producao(dados)

    html = markdown.markdown(
        analise,
        extensions=["tables", "fenced_code"]
    )

    return HttpResponse(html)


def posicao_estoque_view(request):

    nome_estoque = request.GET.get("estoque", "").strip()

    dados = buscar_posicao_estoque(nome_estoque)

    estoques = sorted(
        set(
            item["nome_estoque"]
            for item in dados
            if item.get("nome_estoque")
        )
    )

    contexto = {
        "dados": dados,
        "estoques": estoques,
        "estoque_selecionado": nome_estoque,
    }

    return render(
        request,
        "relatorios/posicao_estoque.html",
        contexto
    )


def exportar_posicao_estoque_excel(request):
    """
    Exporta a posição de estoque para Excel.
    Respeita o filtro de estoque aplicado.
    """

    # Recupera o mesmo filtro utilizado na tela
    nome_estoque = request.GET.get("estoque", "").strip()

    # Busca os dados
    dados = buscar_posicao_estoque(nome_estoque)

    # Converte para DataFrame
    df = pd.DataFrame(dados)

    # Cria resposta Excel
    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="posicao_estoque.xlsx"'
    )

    # Exporta
    df.to_excel(
        response,
        index=False,
        sheet_name="Posição Estoque"
    )

    return response


def relatorios_diversos(request):

    return render(
        request,
        "relatorios/relatorios_diversos.html"
    )
    
    
def previsao_demanda(request):

    status = request.GET.get("Status")
    Temperatura = request.GET.get("Temperatura")
    previsao = request.GET.get("Previsão")

    dados = buscar_previsao_demanda(
        status=status,
        Temperatura=Temperatura,
        previsao=previsao
    )

    # Valores disponíveis para os filtros
    todos_dados = buscar_previsao_demanda()

    status_lista = sorted(
        set(
            item["Status"]
            for item in todos_dados
            if item.get("Status")
        )
    )

    temperatura_lista = sorted(
        set(
            item["Temperatura"]
            for item in todos_dados
            if item.get("Temperatura")
        )
    )

    previsao_lista = sorted(
    set(
            item["Previsão"].strftime("%Y-%m-%d")
            for item in todos_dados
            if item.get("Previsão")
        )
    )

    return render(
        request,
        "relatorios/previsao_demanda.html",
        {
            "dados": dados,
            "status_lista": status_lista,
            "temperatura_lista": temperatura_lista,
            "previsao_lista": previsao_lista,
        }
    )
    
    
def exportar_previsao_demanda_excel(request):

    # ==========================================
    # RECEBE OS FILTROS DA TELA
    # ==========================================

    status = request.GET.get("Status")
    Temperatura = request.GET.get("Temperatura")
    previsao = request.GET.get("Previsão")


    # ==========================================
    # BUSCA OS DADOS USANDO A MESMA FUNÇÃO
    # DO RELATÓRIO
    # ==========================================

    dados = buscar_previsao_demanda(
        status=status,
        Temperatura=Temperatura,
        previsao=previsao
    )


    # ==========================================
    # CRIA O ARQUIVO EXCEL
    # ==========================================

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Previsão de Demanda"


    # ==========================================
    # COLUNAS DA VIEW
    # ==========================================

    colunas = [
        "DataInclusão",
        "Cliente",
        "Previsão",
        "Produto",
        "Temperatura",
        "Fase",
        "Status",
        "Valor",
        "SLA",
    ]


    # ==========================================
    # CABEÇALHO
    # ==========================================

    for numero_coluna, nome_coluna in enumerate(colunas, start=1):

        celula = worksheet.cell(
            row=1,
            column=numero_coluna,
            value=nome_coluna
        )

        celula.font = Font(
            bold=True,
            color="FFFFFF"
        )

        celula.fill = PatternFill(
            fill_type="solid",
            fgColor="0D3B66"
        )

        celula.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )


    # ==========================================
    # DADOS
    # ==========================================

    for numero_linha, item in enumerate(dados, start=2):

        for numero_coluna, nome_coluna in enumerate(colunas, start=1):

            valor = item.get(nome_coluna)

            celula = worksheet.cell(
                row=numero_linha,
                column=numero_coluna,
                value=valor
            )

            celula.alignment = Alignment(
                vertical="top"
            )


    # ==========================================
    # FORMATAÇÃO DAS DATAS
    # ==========================================

    for linha in range(2, worksheet.max_row + 1):

        worksheet.cell(
            row=linha,
            column=1
        ).number_format = "dd/mm/yyyy"


        worksheet.cell(
            row=linha,
            column=3
        ).number_format = "dd/mm/yyyy"


    # ==========================================
    # LARGURA DAS COLUNAS
    # ==========================================

    larguras = {
        1: 16,   # DataInclusão
        2: 40,   # Cliente
        3: 16,   # Previsão
        4: 40,   # Produto
        5: 16,   # Temperatura
        6: 25,   # Fase
        7: 16,   # Status
        8: 18,   # Valor
        9: 16,   # SLA
    }


    for numero_coluna, largura in larguras.items():

        worksheet.column_dimensions[
            get_column_letter(numero_coluna)
        ].width = largura


    # ==========================================
    # CONGELA O CABEÇALHO
    # ==========================================

    worksheet.freeze_panes = "A2"


    # ==========================================
    # FILTRO AUTOMÁTICO DO EXCEL
    # ==========================================

    if worksheet.max_row > 1:

        worksheet.auto_filter.ref = worksheet.dimensions


    # ==========================================
    # RESPOSTA PARA DOWNLOAD
    # ==========================================

    resposta = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


    data_atual = datetime.now().strftime("%Y-%m-%d")


    resposta["Content-Disposition"] = (
        f'attachment; filename="previsao_demanda_{data_atual}.xlsx"'
    )


    workbook.save(resposta)


    return resposta
    
    
    
    
def qualidade_controle(request):

    ops = listar_ops()

    for op in ops.get("cadastros", []):

        codigo_produto = op["identificacao"]["nCodProduto"]

        produto = consultar_produto(codigo_produto)

        op["nome_produto"] = produto.get("descricao", "Produto não encontrado")
        
        
        op["codigo_produto"] = produto.get(
            "codigo",
            ""
        )

        op["peso_liq"] = produto.get(
            "peso_liq"
        )

        op["peso_bruto"] = produto.get(
            "peso_bruto"
        )

    return render(
        request,
        "controle_ops.html",
        {
            "ops": ops
        }
    )

    

from decimal import Decimal


def controle_peso(request):

    # ==========================================================
    # SALVAR CONTROLE DE QUALIDADE
    # ==========================================================

    if request.method == "POST":

        try:

            dados = json.loads(request.body)

            codigo_op = dados.get("codigo_op")
            codigo_produto = dados.get("codigo_produto")
            produto = dados.get("produto")

            data_controle = dados.get("data_controle")
            responsavel = dados.get("responsavel")

            peso_nominal = dados.get("peso_nominal")
            peso_minimo = dados.get("peso_minimo")
            peso_maximo = dados.get("peso_maximo")

            observacao = dados.get("observacao")

            inspecoes = dados.get("inspecoes", [])

            ids_salvos = []


            # ==================================================
            # SALVA CADA INSPEÇÃO
            # ==================================================

            for inspecao in inspecoes:

                controle_id = salvar_controle_qualidade(

                    codigo_op=codigo_op,
                    codigo_produto=codigo_produto,
                    produto=produto,

                    data_controle=data_controle,
                    responsavel=responsavel,

                    peso_nominal=peso_nominal,
                    peso_minimo=peso_minimo,
                    peso_maximo=peso_maximo,

                    numero_inspecao=inspecao.get(
                        "numero_inspecao"
                    ),

                    horario=inspecao.get(
                        "horario"
                    ),

                    media_peso=inspecao.get(
                        "media_peso"
                    ),

                    embalagem=inspecao.get(
                        "embalagem"
                    ),

                    selagem=inspecao.get(
                        "selagem"
                    ),

                    rotulagem=inspecao.get(
                        "rotulagem"
                    ),

                    codificacao=inspecao.get(
                        "codificacao"
                    ),

                    lacre=inspecao.get(
                        "lacre"
                    ),

                    resultado=inspecao.get(
                        "resultado"
                    ),

                    motivo=inspecao.get(
                        "motivo"
                    ),

                    observacao=observacao,

                    pesagens=inspecao.get(
                        "pesagens",
                        []
                    )
                )

                ids_salvos.append(controle_id)


            return JsonResponse({
                "sucesso": True,
                "ids": ids_salvos
            })


        except Exception as e:

            print(
                "❌ ERRO AO SALVAR CONTROLE DE QUALIDADE:"
            )

            print(e)

            return JsonResponse(
                {
                    "sucesso": False,
                    "erro": str(e)
                },
                status=500
            )


    # ==========================================================
    # GET — ABRIR CONTROLE DE PESO
    # ==========================================================

    numero_op = request.GET.get("op")

    ops = listar_ops()

    for op in ops.get("cadastros", []):

        if op["identificacao"]["cNumOP"] == numero_op:

            codigo_produto = op["identificacao"]["nCodProduto"]

            produto = consultar_produto(codigo_produto)

            op["nome_produto"] = produto.get(
                "descricao",
                "Produto não encontrado"
            )

            op["codigo_produto"] = produto.get(
                "codigo",
                ""
            )

            print(
                f"Produto: {op['nome_produto']}, "
                f"Código: {op['codigo_produto']}"
            )


            # ==========================================
            # PESOS DO PRODUTO
            # Omie retorna em KG
            # ==========================================

            peso_liq = produto.get("peso_liq") or 0
            peso_bruto = produto.get("peso_bruto") or 0


            # Converter KG -> GRAMAS

            op["peso_liq_g"] = (
                float(peso_liq) * 1000
            )

            op["tara_g"] = (
                float(peso_bruto) * 1000
            )


            # ==========================================
            # TOLERÂNCIA DE PESAGEM
            # ==========================================

            tolerancia = 0.01


            op["peso_minimo_g"] = (
                op["peso_liq_g"] *
                (1 - tolerancia)
            )

            op["peso_maximo_g"] = (
                op["peso_liq_g"] *
                (1 + tolerancia)
            )


            # ==========================================
            # ABRE A PÁGINA
            # ==========================================

            return render(
                request,
                "controle_peso.html",
                {
                    "op": op
                }
            )


    # ==========================================================
    # OP NÃO ENCONTRADA
    # ==========================================================

    return render(
        request,
        "controle_peso.html",
        {
            "op": None
        }
    )
    

def chao_fabrica(request):

    dados = consulta_chao_fabrica()

    return render(
        request,
        "relatorios/chao_fabrica.html",
        {
            "dados": dados
        }
    )
    
    
def atualizar_ordem_chao_fabrica(request):

    if request.method != "POST":

        return JsonResponse(
            {
                "sucesso": False,
                "erro": "Método não permitido."
            },
            status=405
        )

    try:

        dados = json.loads(
            request.body
        )

        salvar_ordem_chao_fabrica(
            dados
        )

        return JsonResponse(
            {
                "sucesso": True
            }
        )

    except Exception as e:

        return JsonResponse(
            {
                "sucesso": False,
                "erro": str(e)
            },
            status=500
        )